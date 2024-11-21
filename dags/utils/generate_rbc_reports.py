import logging
import os
import urllib.parse
from datetime import date, datetime
from enum import Enum
from typing import Tuple

import numpy as np
import pandas as pd
import sqlalchemy as sa
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class Configuracion:
    """Configuración general del programa"""

    DIRECTORIO_LOGS = "logs"
    DIRECTORIO_REPORTES = "Reportes"
    DIRECTORIO_IMAGENES = "Imagenes"
    NOMBRE_LOGO = "Intelica-logo.png"
    FORMATO_FECHA = "%Y-%m-%d"
    FORMATO_MES = "%Y-%m"
    RUTA_LOGO = os.path.join(DIRECTORIO_IMAGENES, NOMBRE_LOGO)


class ConfiguracionDB:
    """Configuración de la base de datos"""

    SERVIDOR = "10.0.4.100"
    BASE_DATOS = "ITLRPT"
    USUARIO = "SoporteITX"
    CONTRASEÑA = "kSN^3xrW7Mw0"
    DRIVER = "ODBC Driver 17 for SQL Server"

    @classmethod
    def obtener_cadena_conexion(cls) -> str:
        """Generar la cadena de conexión para SQL Server"""
        cadena_conexion = (
            f"Driver={{{cls.DRIVER}}};"
            f"Server={cls.SERVIDOR};"
            f"Database={cls.BASE_DATOS};"
            f"UID={cls.USUARIO};"
            f"PWD={cls.CONTRASEÑA};"
        )
        return urllib.parse.quote_plus(cadena_conexion)


class TipoReporte(Enum):
    DEBITO = "DEBIT"
    CREDITO = "CREDIT"


def configurar_logging() -> logging.Logger:
    """Configura el sistema de logging"""
    log_dir = Configuracion.DIRECTORIO_LOGS
    os.makedirs(log_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_dir, f"reporte_interchange_{timestamp}.log")

    # Formato de logging corregido - usando solo los campos disponibles
    formato = logging.Formatter(
        "%(asctime)s - %(levelname)s - %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    )

    # Handler para archivo
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(formato)

    # Handler para consola
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formato)

    # Configurar logger
    logger = logging.getLogger(
        "reporte_interchange"
    )  # Nombre específico para el logger
    logger.setLevel(logging.INFO)

    # Remover handlers existentes si hay
    if logger.handlers:
        logger.handlers.clear()

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


def obtener_fechas_mes_anterior() -> Tuple[str, str]:
    """Obtener el primer y último día del mes anterior"""
    hoy = date.today()
    primer_dia = hoy.replace(day=1) - relativedelta(months=1)
    ultimo_dia = hoy.replace(day=1) - relativedelta(days=1)
    return primer_dia.strftime(Configuracion.FORMATO_FECHA), ultimo_dia.strftime(
        Configuracion.FORMATO_FECHA
    )


def crear_motor_conexion() -> sa.engine.Engine:
    """Crear motor SQLAlchemy con parámetros de conexión"""
    params = ConfiguracionDB.obtener_cadena_conexion()
    return sa.create_engine(
        f"mssql+pyodbc:///?odbc_connect={params}", fast_executemany=True
    )


def validar_dataframe(df: pd.DataFrame) -> bool:
    """Validar que el DataFrame tenga las columnas necesarias"""
    columnas_requeridas = ["SettlementDate", "Country"]
    return all(columna in df.columns for columna in columnas_requeridas)


def obtener_datos_reporte(
    motor: sa.engine.Engine, tipo_reporte: TipoReporte, logger: logging.Logger
) -> pd.DataFrame:
    """Obtener datos de la base de datos según el tipo de reporte"""
    try:
        fecha_inicio, fecha_fin = obtener_fechas_mes_anterior()
        nombre_tabla = f"ITLRPT.DBO.RPT_ITX_RBC_ISS_{tipo_reporte.value}"

        query = f"""
        SELECT * FROM {nombre_tabla}
        WHERE SettlementDate BETWEEN '{fecha_inicio}' AND '{fecha_fin}'
        """

        logger.info(f"Obteniendo datos del reporte {tipo_reporte.value}")
        inicio = datetime.now()

        df = pd.read_sql_query(query, motor)

        if not validar_dataframe(df):
            raise ValueError("El DataFrame no contiene las columnas requeridas")

        if df.empty:
            raise ValueError(
                f"No se encontraron datos para el período {fecha_inicio} a {fecha_fin}"
            )

        fin = datetime.now()
        logger.info(f"Tiempo de ejecución: {fin - inicio}")
        return df

    except Exception as e:
        logger.error(f"Error al obtener datos: {str(e)}")
        raise


def reemplazar_nulos_por_ceros(df: pd.DataFrame) -> pd.DataFrame:
    """Reemplazar valores nulos por ceros en columnas numéricas"""
    df_copy = df.copy()
    columnas_numericas = df_copy.select_dtypes(include=[np.number]).columns
    df_copy[columnas_numericas] = df_copy[columnas_numericas].fillna(0)
    return df_copy


def aplicar_formato_hoja(
    df: pd.DataFrame, pais: str, tipo_reporte: TipoReporte, logger: logging.Logger
):
    """Aplicar formato a la hoja de Excel"""
    try:
        mes_año = (
            pd.to_datetime(df["SettlementDate"])
            .dt.strftime(Configuracion.FORMATO_MES)
            .unique()[0]
        )
        directorio_base = os.path.join(
            Configuracion.DIRECTORIO_REPORTES, tipo_reporte.value.lower(), mes_año
        )
        os.makedirs(directorio_base, exist_ok=True)

        archivo = os.path.join(
            directorio_base,
            f"Interchange Report Issuer {tipo_reporte.value} {mes_año}.xlsx",
        )

        with pd.ExcelWriter(
            archivo, engine="openpyxl", mode="a" if os.path.exists(archivo) else "w"
        ) as writer:
            if not hasattr(writer, "book"):
                writer.book = (
                    load_workbook(archivo) if os.path.exists(archivo) else Workbook()
                )

            if pais in writer.book.sheetnames:
                del writer.book[pais]

            # Escribir datos
            df.to_excel(writer, sheet_name=pais, index=False, startcol=1, startrow=12)
            hoja = writer.book[pais]

            # Formato de fechas y título
            fecha_inicio = (
                pd.to_datetime(df["SettlementDate"])
                .dt.strftime(Configuracion.FORMATO_FECHA)
                .min()
            )
            fecha_fin = (
                pd.to_datetime(df["SettlementDate"])
                .dt.strftime(Configuracion.FORMATO_FECHA)
                .max()
            )

            # Aplicar formatos
            celda_B9 = hoja["B9"]
            celda_B9.value = (
                f"Settlement Report From {fecha_inicio} to {fecha_fin} - {pais}"
            )
            celda_B9.font = Font(name="Calibri", bold=True, size=16, color="17375E")

            celda_M12 = hoja["M12"]
            celda_M12.value = "Total"
            celda_M12.font = Font(name="Calibri", bold=True, size=12, color="000000")

            # Formato de encabezados
            for celda in hoja["B13":"V13"][0]:
                celda.font = Font(name="Calibri", bold=True, color="FFFFFF")
                celda.fill = PatternFill(
                    start_color="17375E", end_color="17375E", fill_type="solid"
                )

            # Autoajustar columnas
            for col in hoja.iter_cols(min_col=3, max_col=hoja.max_column):
                max_length = 0
                for celda in col:
                    try:
                        if len(str(celda.value)) > max_length:
                            max_length = len(str(celda.value))
                    except:  # noqa: E722
                        pass
                adjusted_width = (max_length + 2) * 1.2
                hoja.column_dimensions[col[0].column_letter].width = adjusted_width

            # Ocultar líneas de cuadrícula
            hoja.sheet_view.showGridLines = False

            # Agregar logo
            try:
                if os.path.exists(Configuracion.RUTA_LOGO):
                    img = Image(Configuracion.RUTA_LOGO)
                    img.width = 350
                    img.height = 100
                    hoja.add_image(img, "B2")
                else:
                    logger.warning(
                        f"No se encontró el logo en {Configuracion.RUTA_LOGO}"
                    )
            except Exception as e:
                logger.warning(f"Error al agregar el logo: {str(e)}")

            # Aplicar formatos de bordes y totales
            medium = Side(border_style="medium", color="000000")
            thin = Side(border_style="thin", color="000000")

            # Calcular y formatear totales
            for col, columna in enumerate(df.columns, start=1):
                if np.issubdtype(df[columna].dtype, np.number):
                    suma = df[columna].sum()
                    celda = hoja.cell(row=12, column=col + 1)
                    celda.value = suma
                    celda.alignment = Alignment(horizontal="center", vertical="center")
                    celda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    celda.number_format = (
                        "#,##0" if df[columna].dtype == "int64" else "#,##0.0"
                    )

            # Combinar celdas y aplicar formatos especiales
            hoja.merge_cells("N11:P11")
            hoja.merge_cells("Q11:S11")
            hoja.merge_cells("T11:V11")

            # Configurar celdas especiales
            celdas_especiales = {
                "N11": ("Base II Clearing File", "17375E"),
                "Q11": ("VSS Reports", "E66914"),
                "T11": ("VSS Reports - Base II Clearing File", "949494"),
            }

            for pos, (texto, color) in celdas_especiales.items():
                celda = hoja[pos]
                celda.value = texto
                celda.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
                celda.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                celda.border = Border(
                    top=medium, left=medium, right=medium, bottom=thin
                )
                celda.alignment = Alignment(horizontal="center", vertical="center")

            # Aplicar formato numérico a todas las celdas
            for col in hoja.iter_cols(min_col=hoja.min_column, max_col=hoja.max_column):
                for celda in col:
                    if isinstance(celda.value, float):
                        celda.number_format = "#,##0.0"
                    elif isinstance(celda.value, int):
                        celda.number_format = "#,##0"

        logger.info(f"Hoja formateada exitosamente para {pais} en {archivo}")

    except Exception as e:
        logger.error(f"Error al formatear hoja para {pais}: {str(e)}")
        raise


def main():
    """Función principal de ejecución"""
    logger = configurar_logging()
    logger.info("Iniciando proceso de generación de reportes")

    try:
        motor = crear_motor_conexion()

        for tipo_reporte in TipoReporte:
            try:
                logger.info(f"Iniciando procesamiento de {tipo_reporte.value}")
                df = obtener_datos_reporte(motor, tipo_reporte, logger)

                if df.empty:
                    logger.warning(
                        f"No hay datos para procesar en {tipo_reporte.value}"
                    )
                    continue

                df_formateado = reemplazar_nulos_por_ceros(df)
                paises = df_formateado["Country"].dropna().unique()

                for pais in paises:
                    try:
                        logger.info(f"Procesando {pais}")
                        df_pais = df_formateado[df_formateado["Country"] == pais]
                        aplicar_formato_hoja(df_pais, pais, tipo_reporte, logger)
                    except Exception as e:
                        logger.error(f"Error procesando {pais}: {str(e)}")
                        continue

                logger.info(f"Completado procesamiento de {tipo_reporte.value}")

            except Exception as e:
                logger.error(
                    f"Error en el procesamiento de {tipo_reporte.value}: {str(e)}"
                )
                continue

        logger.info("Proceso completado exitosamente")

    except Exception as e:
        logger.error(f"Error fatal en la ejecución: {str(e)}")
        raise
